import csv
import os
import re
import sys
import tempfile
from typing import Iterable, List, Optional, Sequence, Tuple

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Optional: Silbentrennung (nur Header)
try:
    import pyphen
except Exception:
    pyphen = None

# =========================
# Konstanten & Presets
# =========================
DPI = 300
A4_MM = (210, 297)

HEADER_HARD_WRAP_CHARS = 5   # Header „hart“ umbrechen
BODY_HARD_WRAP_CHARS = 18    # Body „harte“ Chunks (lange Tokens, z. B. Suchstrings)

# Obergrenzen für Spaltenanteile an der Seitenbreite
MAX_COL_SHARE_DEFAULT = 0.45     # allgemein
MAX_COL_SHARE_QUERY = 0.60       # Spalten, die wie Suchqueries aussehen

COLOR_PRESETS = [
    ("Hellgrau", "#F5F5F5"),
    ("Wolkenweiß", "#FFFFFF"),
    ("Blassblau", "#EEF4FB"),
    ("Eisblau", "#F7FBFF"),
    ("Mint", "#ECF7F2"),
    ("Salbei", "#F1F7F2"),
    ("Lavendel", "#F3F0FA"),
    ("Altrosa", "#FDF0F3"),
    ("Sand", "#F6F1EB"),
]
PRESET_NAMES = [name for name, _ in COLOR_PRESETS] + ["Benutzerdefiniert…"]
NAME_TO_HEX = {name: hex_value for name, hex_value in COLOR_PRESETS}

HYPHENATOR: Optional["pyphen.Pyphen"] = None
HYPHENATE_HEADERS = False


def configure_hyphenation(enable_headers: bool, hyphenator: Optional["pyphen.Pyphen"]) -> None:
    """Globales Setzen der Silbentrennung (nur Header)."""
    global HYPHENATOR, HYPHENATE_HEADERS
    HYPHENATOR = hyphenator
    HYPHENATE_HEADERS = enable_headers


def resource_path(relative_path: str) -> str:
    """PyInstaller-kompatibler Ressourcenpfad."""
    try:
        base_path = sys._MEIPASS  # type: ignore[attr-defined]
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def mm_to_px(mm: float, dpi: int = DPI) -> int:
    return int(round(mm / 25.4 * dpi))


def a4_pixels(orientation: str = "portrait", dpi: int = DPI) -> Tuple[int, int]:
    w_mm, h_mm = A4_MM
    if orientation == "landscape":
        w_mm, h_mm = h_mm, w_mm
    return mm_to_px(w_mm, dpi), mm_to_px(h_mm, dpi)


def hex_to_rgb(hex_value: str) -> Tuple[int, int, int]:
    try:
        cleaned = hex_value.strip()
        if cleaned.startswith("#"):
            cleaned = cleaned[1:]
        if len(cleaned) == 3:
            cleaned = "".join(ch * 2 for ch in cleaned)
        return tuple(int(cleaned[i:i + 2], 16) for i in (0, 2, 4))
    except Exception:
        return 245, 245, 245


def _choose_existing(paths: Iterable[str]) -> Optional[str]:
    return next((p for p in paths if os.path.exists(p)), None)


def load_fonts(size_body: int = 24, size_header: int = 28) -> Tuple[ImageFont.ImageFont, ImageFont.ImageFont]:
    candidates = [
        (r"C:\Windows\Fonts\segoeui.ttf", r"C:\Windows\Fonts\segoeuib.ttf"),
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf"),
        (r"C:\Windows\Fonts\calibri.ttf", r"C:\Windows\Fonts\calibrib.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/System/Library/Fonts/Supplemental/Arial.ttf",
         "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ("/System/Library/Fonts/Helvetica.ttc", "/System/Library/Fonts/Helvetica.ttc"),
    ]
    for normal, bold in candidates:
        normal_path = _choose_existing([normal])
        bold_path = _choose_existing([bold])
        if normal_path and bold_path:
            try:
                return (
                    ImageFont.truetype(normal_path, size_body),
                    ImageFont.truetype(bold_path, size_header),
                )
            except Exception:
                continue
    # Fallback
    return ImageFont.load_default(), ImageFont.load_default()


def text_size(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> Tuple[int, int]:
    bbox = draw.textbbox((0, 0), str(text), font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int,
              hard_chunk: Optional[int] = None) -> List[str]:
    """Zeilenumbruch: Wortweise, optional mit 'hartem' Chunk-Umbruch für lange 'Wörter'."""
    content = str(text or "")
    if not content:
        return [""]
    inner = max(1, max_width)
    use_hyphenation = bool(hard_chunk) and HYPHENATE_HEADERS and HYPHENATOR is not None

    lines: List[str] = []
    current = ""

    def text_width(fragment: str) -> int:
        return text_size(draw, fragment, font)[0]

    def hyphen_split_once(token: str) -> Tuple[Optional[str], str]:
        if not use_hyphenation or not token:
            return None, token
        try:
            syllables = (HYPHENATOR.inserted(token) or token).split("\u00AD")  # type: ignore[arg-type]
            if len(syllables) <= 1:
                return None, token
            prefix = ""
            consumed = 0
            for idx in range(len(syllables) - 1):
                candidate = prefix + syllables[idx]
                if text_width(candidate + "-") <= inner:
                    prefix = candidate
                    consumed = idx + 1
                else:
                    break
            if consumed > 0:
                return prefix + "\u00AD", "".join(syllables[consumed:])
            return None, token
        except Exception:
            return None, token

    def split_token(token: str) -> List[str]:
        parts: List[str] = []
        remaining = token
        while remaining:
            if text_width(remaining) <= inner:
                parts.append(remaining)
                break
            prefix, rest = hyphen_split_once(remaining)
            if prefix is not None:
                parts.append(prefix)
                remaining = rest
                continue
            limit = min(len(remaining), hard_chunk) if hard_chunk and hard_chunk > 0 else len(remaining)
            lo, hi, fit = 1, limit, 0
            while lo <= hi:
                mid = (lo + hi) // 2
                if text_width(remaining[:mid] + "-") <= inner:
                    fit = mid
                    lo = mid + 1
                else:
                    hi = mid - 1
            if fit == 0:
                fit = 1
            parts.append(remaining[:fit] + "\u00AD")
            remaining = remaining[fit:]
        return parts

    for token in content.split(" "):
        token = token or ""
        trial = (current + " " + token) if current else token
        if current and text_width(trial) <= inner:
            current = trial
            continue
        if current:
            lines.append(current)
            current = ""
        for idx, piece in enumerate(split_token(token)):
            separator = " " if (idx == 0 and current) else ""
            trial = (current + separator + piece) if current else piece
            if current and text_width(trial) > inner:
                lines.append(current)
                current = piece
            else:
                current = trial
            if piece.endswith("\u00AD"):
                lines.append(current)
                current = ""
    if current:
        lines.append(current)
    return lines


def ellipsize(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> str:
    value = str(text or "")
    if not value:
        return ""
    if text_size(draw, value, font)[0] <= max_width:
        return value
    ellipsis = "…"
    lo, hi = 0, len(value)
    while lo < hi:
        mid = (lo + hi) // 2
        candidate = value[:mid] + ellipsis
        if text_size(draw, candidate, font)[0] <= max_width:
            lo = mid + 1
        else:
            hi = mid
    return (value[:max(lo - 1, 0)] + ellipsis) if lo > 0 else ellipsis


# =========================
# CSV / TXT
# =========================
def read_csv(path: str, delimiter: str, encoding_utf8: bool = True) -> List[List[str]]:
    encoding = "utf-8-sig" if encoding_utf8 else "cp1252"
    rows: List[List[str]] = []
    with open(path, "r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows.extend(reader)
    return rows


def normalize_rows(rows: Sequence[Sequence[str]]) -> List[List[str]]:
    if not rows:
        return []
    cleaned: List[List[str]] = []
    for row in rows:
        if row is None:
            continue
        normalized = [("" if cell is None else str(cell)) for cell in row]
        if any(fragment.strip() for fragment in normalized):
            cleaned.append(normalized)
    if not cleaned:
        return []
    if cleaned[0] and isinstance(cleaned[0][0], str):
        cleaned[0][0] = cleaned[0][0].lstrip("\ufeff")
    max_cols = max(len(r) for r in cleaned)
    return [(r + [""] * (max_cols - len(r))) if len(r) < max_cols else list(r[:max_cols]) for r in cleaned]


def create_temp_csv_without_empty_columns(path: str, delimiter: str, encoding_utf8: bool = True) -> str:
    rows = normalize_rows(read_csv(path, delimiter, encoding_utf8=encoding_utf8))
    if not rows:
        raise ValueError("Die CSV-Datei enthält keine verwertbaren Daten.")
    max_cols = max(len(r) for r in rows)
    padded = [(r + [""] * (max_cols - len(r))) if len(r) < max_cols else r[:max_cols] for r in rows]
    header, body = padded[0], padded[1:]
    keep_indices = [idx for idx in range(max_cols) if any(str(row[idx]).strip() for row in body)]
    if not keep_indices:
        raise ValueError("Alle Spalten sind im Body leer – nichts zu behalten.")
    fd, temp_path = tempfile.mkstemp(prefix="csv_clean_", suffix=".csv")
    os.close(fd)
    encoding = "utf-8-sig" if encoding_utf8 else "cp1252"
    with open(temp_path, "w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        for row in padded:
            writer.writerow([row[idx] for idx in keep_indices])
    return temp_path


def create_temp_csv_without_selected_columns(
    path: str,
    delimiter: str,
    columns_to_remove_1based: Sequence[int],
    encoding_utf8: bool = True
) -> str:
    rows = normalize_rows(read_csv(path, delimiter, encoding_utf8=encoding_utf8))
    if not rows:
        raise ValueError("Die CSV-Datei enthält keine verwertbaren Daten.")

    max_cols = max(len(r) for r in rows)
    padded = [(r + [""] * (max_cols - len(r))) if len(r) < max_cols else r[:max_cols] for r in rows]

    invalid_1based = sorted({int(n) for n in columns_to_remove_1based if int(n) < 1 or int(n) > max_cols})
    if invalid_1based:
        raise ValueError(f"Ungültige Spaltennummer(n): {invalid_1based} — die Datei hat nur {max_cols} Spalten.")

    to_remove = set()
    for num in columns_to_remove_1based:
        try:
            idx0 = int(num) - 1
        except Exception:
            continue
        if 0 <= idx0 < max_cols:
            to_remove.add(idx0)

    if not to_remove:
        fd, temp_path = tempfile.mkstemp(prefix="csv_keep_", suffix=".csv")
        os.close(fd)
        encoding = "utf-8-sig" if encoding_utf8 else "cp1252"
        with open(temp_path, "w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter)
            writer.writerows(padded)
        return temp_path

    keep_indices = [idx for idx in range(max_cols) if idx not in to_remove]
    if not keep_indices:
        raise ValueError("Alle Spalten wurden zum Entfernen ausgewählt — es bleibt nichts übrig.")

    fd, temp_path = tempfile.mkstemp(prefix="csv_cols_removed_", suffix=".csv")
    os.close(fd)
    encoding = "utf-8-sig" if encoding_utf8 else "cp1252"
    with open(temp_path, "w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        for row in padded:
            writer.writerow([row[idx] for idx in keep_indices])

    return temp_path


def read_cochrane_txt(path: str) -> Tuple[List[str], List[List[str]], dict]:
    text = _read_text_with_fallback(path)
    return _parse_cochrane_search_manager_txt(text)


def _read_text_with_fallback(path: str) -> str:
    for encoding in ("utf-8", "utf-16", "utf-16-le", "utf-16-be", "latin-1", "cp1252"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                return handle.read()
        except Exception:
            continue
    with open(path, "rb") as handle:
        return handle.read().decode("utf-8", errors="ignore")


def _parse_cochrane_search_manager_txt(text: str) -> Tuple[List[str], List[List[str]], dict]:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    meta: dict = {}
    for key in ("Search Name", "Date Run", "Comment"):
        match = re.search(rf"(?m)^{re.escape(key)}:\t(.*)$", text)
        if match:
            meta[key] = match.group(1).strip()
    header_match = re.search(r"(?m)^ID\s*\t\s*Search\s*\t\s*Hits\s*$", text)
    if not header_match:
        raise ValueError("Konnte den Tabellenkopf 'ID\\tSearch\\tHits' nicht finden.")
    lines = text[header_match.end():].strip("\n").split("\n")
    header = ["ID", "Search", "Hits"]
    rows: List[List[str]] = []
    current_id: Optional[str] = None
    buffer: List[str] = []

    def flush_row(final_chunk: Optional[str], hits_str: str) -> None:
        nonlocal rows, buffer, current_id
        collapsed = re.sub(r"\s+", " ", " ".join(p.replace("\t", " ").strip() for p in (buffer + ([final_chunk] if final_chunk is not None else [])) if p)).strip()
        disp_id = f"#{current_id}" if current_id and not str(current_id).startswith("#") else (current_id or "")
        rows.append([disp_id, collapsed, hits_str.strip()])
        buffer.clear()
        current_id = None

    id_line_re = re.compile(r"^#?(\d+)\t(.*)$")
    for line in lines:
        if not line.strip():
            continue
        match = id_line_re.match(line)
        if match and current_id is None:
            current_id = match.group(1)
            rest = match.group(2)
            hits_match = re.search(r"\t(\d+)\s*$", rest)
            if hits_match:
                flush_row(rest[:hits_match.start()], hits_match.group(1))
            else:
                buffer = [rest]
            continue
        if current_id is not None:
            hits_match = re.search(r"\t(\d+)\s*$", line)
            if hits_match:
                flush_row(line[:hits_match.start()], hits_match.group(1))
            else:
                buffer.append(line)
            continue
        if match:
            current_id = match.group(1)
            rest = match.group(2)
            hits_match = re.search(r"\t(\d+)\s*$", rest)
            if hits_match:
                flush_row(rest[:hits_match.start()], hits_match.group(1))
            else:
                buffer = [rest]
    if current_id is not None and buffer:
        flush_row(" ", "")
    return header, rows, meta


# =========================
# Layout / Breiten
# =========================
def _guess_lang_from_header(header_list: Sequence[str]) -> str:
    text = " ".join(map(str, header_list or []))
    if any(ch in text for ch in "äöüÄÖÜß"):
        return "de_DE"
    return "en_US"


def make_hyphenator(choice: str, header_for_auto: Sequence[str]) -> Optional["pyphen.Pyphen"]:
    if pyphen is None:
        return None
    if choice.startswith("Aus"):
        return None
    if choice.startswith("Auto"):
        lang = _guess_lang_from_header(header_for_auto)
    elif "de_DE" in choice or "Deutsch" in choice:
        lang = "de_DE"
    else:
        lang = "en_US"
    try:
        return pyphen.Pyphen(lang=lang)  # type: ignore[call-arg]
    except Exception:
        return None


def measure_header_and_words(draw: ImageDraw.ImageDraw, header: Sequence[str], rows: Sequence[Sequence[str]],
                             font_header: ImageFont.ImageFont, font_body: ImageFont.ImageFont, pad_x: int
                             ) -> Tuple[List[int], List[int], List[int]]:
    n = len(header)
    word_re = re.compile(r"\S+")
    header_piece_longest = [0] * n
    body_longest_word = [0] * n
    natural = [0] * n

    for j in range(n):
        header_text = str(header[j])
        natural[j] = max(natural[j], text_size(draw, header_text, font_header)[0] + 2 * pad_x)
        for token in word_re.findall(header_text):
            if HEADER_HARD_WRAP_CHARS and len(token) > HEADER_HARD_WRAP_CHARS:
                for idx in range(0, len(token), HEADER_HARD_WRAP_CHARS):
                    piece = token[idx:idx + HEADER_HARD_WRAP_CHARS]
                    width, _ = text_size(draw, piece, font_header)
                    header_piece_longest[j] = max(header_piece_longest[j], width + 2 * pad_x)
            else:
                width, _ = text_size(draw, token, font_header)
                header_piece_longest[j] = max(header_piece_longest[j], width + 2 * pad_x)

    for row in rows:
        for j in range(min(len(row), n)):
            text = str(row[j])
            natural[j] = max(natural[j], text_size(draw, text, font_body)[0] + 2 * pad_x)
            for token in word_re.findall(text):
                width, _ = text_size(draw, token, font_body)
                body_longest_word[j] = max(body_longest_word[j], width + 2 * pad_x)
    return header_piece_longest, body_longest_word, natural


def min_floor_from_font(draw: ImageDraw.ImageDraw, font_body: ImageFont.ImageFont, pad_x: int,
                        min_chars: int = 3) -> int:
    width, _ = text_size(draw, "W" * min_chars, font_body)
    return width + 2 * pad_x


def compute_wrap_score(draw: ImageDraw.ImageDraw, col_texts: Sequence[str], font: ImageFont.ImageFont,
                       col_width: int, pad_x: int, line_gap: int, is_header: bool = False,
                       body_hard_chunk: Optional[int] = BODY_HARD_WRAP_CHARS) -> int:
    inner = max(1, col_width - 2 * pad_x)
    score = 0
    _, one_h = text_size(draw, "Ag", font)
    for text in col_texts:
        lines = wrap_text(
            draw, str(text), font, inner,
            hard_chunk=HEADER_HARD_WRAP_CHARS if is_header else body_hard_chunk
        )
        score += max(0, len(lines) - 1) * one_h
    return score


def looks_like_query_column(header_text: str, col_texts: Sequence[str], sample: int = 30) -> bool:
    pattern = re.compile(r"\b(AND|OR|NOT|NEAR/?\d*|ADJ\d*|N\d+|W\d+|TI:|AB:|MH:|MeSH|\".+?\"|\[tiab\])\b", re.IGNORECASE)
    score = 0
    if pattern.search(str(header_text or "")):
        score += 1
    for text in list(col_texts)[:sample]:
        t = str(text or "")
        if pattern.search(t):
            score += 2
        score += t.count("(") + t.count(")") + t.count('"')
        if score >= 4:
            return True
    return score >= 3


def flex_fit_widths(base: Sequence[int], minw: Sequence[int], target_width: int) -> List[int]:
    n = len(base)
    if n == 0:
        return []
    total = sum(base) or 1
    widths = [max(minw[i], int(round(base[i] * (target_width / total)))) for i in range(n)]
    diff = target_width - sum(widths)
    order = sorted(range(n), key=lambda idx: base[idx], reverse=True) or list(range(n))
    guard = 0
    while diff != 0 and guard < 10000:
        changed = False
        for idx in order:
            if diff > 0:
                widths[idx] += 1
                diff -= 1
                changed = True
                if diff == 0:
                    break
            else:
                if widths[idx] > minw[idx]:
                    widths[idx] -= 1
                    diff += 1
                    changed = True
                    if diff == 0:
                        break
        if not changed:
            break
        guard += 1
    leftover = target_width - sum(widths)
    if leftover != 0:
        widths[-1] = max(minw[-1], widths[-1] + leftover)
    return widths


def equal_width_with_buffer(place_w: int, header: Sequence[str], rows: Sequence[Sequence[str]],
                            fonts: Tuple[ImageFont.ImageFont, ImageFont.ImageFont],
                            pad_x: int, line_gap: int, max_col_share: float = MAX_COL_SHARE_DEFAULT) -> List[int]:
    """
    Basis wie bisher, ABER die zusätzliche Breite geht bevorzugt in Spalten,
    die wie Suchqueries aussehen; diese Query-Spalten werden untereinander
    gleichbreit gemacht (keine „eine mega-breit, eine schmal“).
    """
    font_body, font_header = fonts
    probe = Image.new("RGB", (100, 100), "white")
    drawer = ImageDraw.Draw(probe)

    header_piece_longest, _, natural = measure_header_and_words(drawer, header, rows, font_header, font_body, pad_x)
    min_floor = min_floor_from_font(drawer, font_body, pad_x, min_chars=3)

    n = len(header)
    if n == 0:
        return []

    minw = [max(min_floor, header_piece_longest[j]) for j in range(n)]
    equal = max(1, place_w // n)
    widths = [max(minw[j], min(natural[j], equal)) for j in range(n)]

    total = sum(widths)
    if total > place_w or sum(minw) > place_w:
        return flex_fit_widths(natural, minw, place_w)

    columns = list(zip(*rows)) if rows else [[] for _ in range(n)]
    is_query_col = [
        looks_like_query_column(header[j] if j < len(header) else "", columns[j] if j < len(columns) else [])
        for j in range(n)
    ]

    buffer_width = place_w - sum(widths)
    if buffer_width <= 0:
        return widths

    caps = [int(place_w * (MAX_COL_SHARE_QUERY if is_query_col[j] else max_col_share)) for j in range(n)]

    q_idx = [j for j in range(n) if is_query_col[j]]
    if not q_idx:
        # klassisch: nach Wrap-Score verteilen
        scores = []
        for j in range(n):
            header_only = [header[j]] if j < len(header) else []
            column_texts = columns[j] if j < len(columns) else []
            score = compute_wrap_score(drawer, header_only, font_header, max(widths[j], 1), pad_x, line_gap, is_header=True)
            score += compute_wrap_score(drawer, column_texts, font_body, max(widths[j], 1), pad_x, line_gap,
                                        is_header=False, body_hard_chunk=BODY_HARD_WRAP_CHARS)
            scores.append(max(score, 1))
        total_score = sum(scores)
        for j in range(n):
            add = int(round(buffer_width * (scores[j] / total_score)))
            widths[j] = min(caps[j], widths[j] + add)
        diff = place_w - sum(widths)
        order = sorted(range(n), key=lambda j: scores[j], reverse=True)
        guard = 0
        while diff != 0 and guard < 10000:
            for j in order:
                if diff == 0:
                    break
                if diff > 0 and widths[j] < caps[j]:
                    widths[j] += 1
                    diff -= 1
                elif diff < 0 and widths[j] > minw[j]:
                    widths[j] -= 1
                    diff += 1
            guard += 1
        return widths

    # Query-Spalten gleichbreit + Puffer dort verteilen
    current_max_q = max(widths[j] for j in q_idx)
    max_cap_q = min(caps[j] for j in q_idx)
    tentative_target = min(max_cap_q, current_max_q + buffer_width // len(q_idx))

    while tentative_target > current_max_q:
        need = sum(max(0, tentative_target - widths[j]) for j in q_idx)
        if need <= buffer_width:
            break
        tentative_target -= 1

    need = sum(max(0, tentative_target - widths[j]) for j in q_idx)
    for j in q_idx:
        inc = max(0, tentative_target - widths[j])
        widths[j] += inc
    buffer_width -= need

    while buffer_width > 0:
        candidates = [j for j in q_idx if widths[j] < caps[j]]
        if not candidates:
            break
        add_each = min(buffer_width // len(candidates), min(caps[j] - widths[j] for j in candidates))
        if add_each <= 0:
            for j in candidates:
                if buffer_width <= 0:
                    break
                if widths[j] < caps[j]:
                    widths[j] += 1
                    buffer_width -= 1
            break
        for j in candidates:
            widths[j] += add_each
        buffer_width -= add_each * len(candidates)

    diff = place_w - sum(widths)
    if diff != 0:
        order = q_idx + [j for j in range(n) if j not in q_idx]
        guard = 0
        while diff != 0 and guard < 10000:
            for j in order:
                if diff == 0:
                    break
                if diff > 0 and widths[j] < caps[j]:
                    widths[j] += 1
                    diff -= 1
                elif diff < 0 and widths[j] > minw[j]:
                    widths[j] -= 1
                    diff += 1
            guard += 1

    return widths


def row_height_for_widths(draw: ImageDraw.ImageDraw, cells: Sequence[str], col_widths: Sequence[int],
                          font: ImageFont.ImageFont, pad_x: int, pad_y: int, line_gap: int,
                          is_header: bool = False) -> int:
    max_lines = 1
    for idx, text in enumerate(cells):
        col_width = max(1, col_widths[idx] - 2 * pad_x)
        lines = wrap_text(
            draw, str(text), font, col_width,
            hard_chunk=HEADER_HARD_WRAP_CHARS if is_header else BODY_HARD_WRAP_CHARS
        )
        max_lines = max(max_lines, len(lines))
    _, one_h = text_size(draw, "Ag", font)
    return one_h * max_lines + 2 * pad_y + (max_lines - 1) * line_gap


# ====== Renderer (PIL) mit Row-Slicing; liefert Seiten als Images ======
def render_pages_dynamic(header: Sequence[str], rows: Sequence[Sequence[str]], orientation: str,
                         header_hex: str, zebra_hex: str, top_note: Optional[str] = None,
                         top_right_text: Optional[str] = None, custom_font_pt: Optional[int] = None) -> List[Image.Image]:
    page_w, page_h = a4_pixels(orientation)
    margin = mm_to_px(12)  # fester Rand (unten = oben = links = rechts)
    place_w, place_h = page_w - 2 * margin, page_h - 2 * margin

    body_size = int(custom_font_pt) if (custom_font_pt and custom_font_pt > 0) else 24
    header_size = max(16, body_size + 4)
    pad_x, pad_y, line_gap = 16, 12, 6

    probe = Image.new("RGB", (100, 100), "white")
    draw0 = ImageDraw.Draw(probe)

    # Header-Schrift ggf. minimal verkleinern, bis Mindestbreiten passen
    min_header = 16
    while True:
        fonts = load_fonts(size_body=body_size, size_header=header_size)
        font_body, font_header = fonts
        n = len(header)
        word_re = re.compile(r"\S+")
        header_piece_longest = [0] * n
        for j in range(n):
            for tok in word_re.findall(str(header[j])):
                if HEADER_HARD_WRAP_CHARS and len(tok) > HEADER_HARD_WRAP_CHARS:
                    for i0 in range(0, len(tok), HEADER_HARD_WRAP_CHARS):
                        piece = tok[i0:i0 + HEADER_HARD_WRAP_CHARS]
                        w, _ = text_size(draw0, piece, font_header)
                        header_piece_longest[j] = max(header_piece_longest[j], w + 2 * pad_x)
                else:
                    w, _ = text_size(draw0, tok, font_header)
                    header_piece_longest[j] = max(header_piece_longest[j], w + 2 * pad_x)
        minw_try = [max(min_floor_from_font(draw0, font_body, pad_x, 3), header_piece_longest[j]) for j in range(n)]
        if sum(minw_try) <= place_w or header_size <= min_header:
            break
        header_size -= 2

    fonts = load_fonts(size_body=body_size, size_header=header_size)
    font_body, font_header = fonts
    col_widths = equal_width_with_buffer(place_w, header, rows, fonts, pad_x, line_gap)

    _, one_h_body = text_size(draw0, "Ag", font_body)
    header_h = row_height_for_widths(draw0, header, col_widths, font_header, pad_x, pad_y, line_gap, is_header=True)
    band_h = (2 * pad_y + one_h_body)

    header_rgb = hex_to_rgb(header_hex)
    zebra_rgb = hex_to_rgb(zebra_hex)
    text_color = (0, 0, 0)
    grid = (200, 200, 200)

    def _wrap_row_cells(row: Sequence[str]) -> List[List[str]]:
        lines_per_col: List[List[str]] = []
        inner_ws = [max(1, w - 2 * pad_x) for w in col_widths]
        for j in range(len(header)):
            txt = "" if j >= len(row) else str(row[j])
            lines = wrap_text(draw0, txt, font_body, inner_ws[j], hard_chunk=BODY_HARD_WRAP_CHARS)
            lines_per_col.append(lines)
        return lines_per_col

    wrapped_rows = [_wrap_row_cells(r) for r in rows]

    def simulate_pages() -> int:
        y = margin + band_h + header_h
        pages = 1 if rows else 1
        for lines_per_col in wrapped_rows:
            offs = [0] * len(lines_per_col)
            while True:
                avail = (margin + place_h) - y
                min_block = (2 * pad_y + one_h_body)
                if avail < min_block:
                    pages += 1
                    y = margin + band_h + header_h
                    continue
                k_fit = max(1, (avail - 2 * pad_y + line_gap) // (one_h_body + line_gap))
                k_used = 0
                for j in range(len(lines_per_col)):
                    k_used = max(k_used, min(k_fit, len(lines_per_col[j]) - offs[j]))
                if k_used <= 0:
                    break
                slice_h = 2 * pad_y + k_used * one_h_body + (k_used - 1) * line_gap
                y += slice_h
                for j in range(len(lines_per_col)):
                    offs[j] = min(len(lines_per_col[j]), offs[j] + min(k_fit, len(lines_per_col[j]) - offs[j]))
                if all(offs[j] >= len(lines_per_col[j]) for j in range(len(lines_per_col))):
                    break
                if y > margin + place_h - min_block:
                    pages += 1
                    y = margin + band_h + header_h
        return max(1, pages)

    total_pages = simulate_pages()

    pages: List[Image.Image] = []
    page_idx = 0

    def start_page(idx: int) -> Tuple[Image.Image, ImageDraw.ImageDraw, int]:
        can = Image.new("RGB", (page_w, page_h), "white")
        dr = ImageDraw.Draw(can)
        y0 = margin
        # Kopfband (oben)
        dr.line([(margin, y0 + band_h - 1), (margin + place_w - 1, y0 + band_h - 1)], fill=grid)
        inner_w = max(1, place_w - 2 * pad_x)
        # links: optionaler Hinweis (z. B. "Date Run")
        left_text = ellipsize(dr, top_note or "", font_body, inner_w // 3) if (top_note and str(top_note).strip()) else ""
        if left_text:
            dr.text((margin + pad_x, y0 + pad_y), left_text, font=font_body, fill=text_color)
        # mitte: Seite i/N
        page_text = f"Seite {idx}/{total_pages}"
        tw, _ = text_size(dr, page_text, font=font_body)
        dr.text((margin + place_w // 2 - tw // 2, y0 + pad_y), page_text, font=font_body, fill=text_color)
        # rechts: freier Kopftext (UI)
        if top_right_text and str(top_right_text).strip():
            rt = ellipsize(dr, top_right_text, font=font_body, max_width=inner_w // 3)
            rtw, _ = text_size(dr, rt, font=font_body)
            dr.text((margin + place_w - pad_x - rtw, y0 + pad_y), rt, font=font_body, fill=text_color)

        # Tabellenkopf-Hintergrund
        y1 = y0 + band_h
        dr.rectangle([margin, y1, margin + place_w - 1, y1 + header_h - 1], fill=header_rgb)
        # Spalten + Headertexte
        x = margin
        for j in range(len(header)):
            w = col_widths[j]
            dr.line([(x, y1), (x, y1 + place_h)], fill=grid)
            inner = max(1, w - 2 * pad_x)
            lines = wrap_text(dr, str(header[j]), font_header, inner, hard_chunk=HEADER_HARD_WRAP_CHARS)
            _, one_hh = text_size(dr, "Ag", font_header)
            total_h = len(lines) * one_hh + (len(lines) - 1) * line_gap
            y_text = y1 + (header_h - total_h) // 2
            for ln in lines:
                dr.text((x + pad_x, y_text), ln, font=font_header, fill=text_color)
                y_text += one_hh + line_gap
            x += w
        # Außenlinien
        dr.line([(margin + place_w - 1, y1), (margin + place_w - 1, y1 + place_h)], fill=grid)
        dr.line([(margin, y1 + header_h - 1), (margin + place_w - 1, y1 + header_h - 1)], fill=grid)
        return can, dr, y1 + header_h

    page_idx += 1
    canvas, drawer, y = start_page(page_idx)

    for r_idx, row in enumerate(rows):
        lines_per_col = []
        inner_ws = [max(1, w - 2 * pad_x) for w in col_widths]
        for j in range(len(header)):
            txt = "" if j >= len(row) else str(row[j])
            lines_per_col.append(wrap_text(draw0, txt, font_body, inner_ws[j], hard_chunk=BODY_HARD_WRAP_CHARS))

        offs = [0] * len(lines_per_col)
        zebra_on = ((r_idx % 2) == 0)

        while True:
            avail = (margin + place_h) - y
            min_block = (2 * pad_y + one_h_body)
            if avail < min_block:
                pages.append(canvas)
                page_idx += 1
                canvas, drawer, y = start_page(page_idx)
                avail = (margin + place_h) - y

            k_fit = max(1, (avail - 2 * pad_y + line_gap) // (one_h_body + line_gap))
            k_used = 0
            per_col_take: List[int] = []
            for j in range(len(lines_per_col)):
                rest = max(0, len(lines_per_col[j]) - offs[j])
                take = min(k_fit, rest)
                per_col_take.append(take)
                k_used = max(k_used, take)

            if k_used <= 0:
                break

            slice_h = 2 * pad_y + k_used * one_h_body + (k_used - 1) * line_gap

            x = margin
            if zebra_on:
                drawer.rectangle([margin, y, margin + place_w - 1, y + slice_h - 1], fill=zebra_rgb)
            for j in range(len(header)):
                w = col_widths[j]
                y_text = y + pad_y
                take = per_col_take[j]
                for ln in lines_per_col[j][offs[j]:offs[j] + take]:
                    drawer.text((x + pad_x, y_text), ln, font=font_body, fill=text_color)
                    y_text += one_h_body + line_gap
                drawer.rectangle([x, y, x + w - 1, y + slice_h - 1], outline=(200, 200, 200))
                offs[j] += take
                x += w

            y += slice_h

            if not all(offs[j] >= len(lines_per_col[j]) for j in range(len(lines_per_col))) and y > (margin + place_h - min_block):
                pages.append(canvas)
                page_idx += 1
                canvas, drawer, y = start_page(page_idx)

    pages.append(canvas)
    return pages


def save_images_as_pdf(pages: List[Image.Image], out_path: str) -> None:
    """Bild-PDF ohne OCR."""
    rgb_pages = [page.convert("RGB") for page in pages]
    if len(rgb_pages) == 1:
        rgb_pages[0].save(out_path, "PDF", resolution=DPI)
    else:
        rgb_pages[0].save(out_path, "PDF", save_all=True, append_images=rgb_pages[1:], resolution=DPI)  # type: ignore[arg-type]


def save_as_excel(header: Sequence[str], rows: Sequence[Sequence[str]], path: str,
                  orientation: str, fit_width: bool = True, meta_note: Optional[str] = None,
                  header_right_text: Optional[str] = None) -> None:
    """XLSX-Ausgabe mit Druckeinstellungen und Kopfzeile."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle"
    first_data_row = 1

    if meta_note:
        ws.append([meta_note])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        ws["A1"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["A1"].font = Font(bold=False, italic=True)
        first_data_row = 2

    ws.append(list(map(str, header)))
    for row in rows:
        ws.append([("" if idx >= len(row) else str(row[idx])) for idx in range(len(header))])

    header_font = Font(bold=True)
    align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row_idx = first_data_row
    for col_idx in range(1, len(header) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws[f"{col_letter}{header_row_idx}"].value or ""))
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            val = ws[f"{col_letter}{row_idx}"].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(80, max(10, int(max_len * 1.2)))

    for row in ws.iter_rows(min_row=header_row_idx, max_row=ws.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.alignment = align
            cell.border = border

    fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for cell in ws[header_row_idx]:
        cell.font = header_font
        cell.fill = fill

    try:
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = "landscape" if orientation == "landscape" else "portrait"
        ws.page_setup.fitToWidth = 1 if fit_width else 0
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = f"{header_row_idx}:{header_row_idx}"
        ws.page_margins.left = 0.39
        ws.page_margins.right = 0.39
        ws.page_margins.top = 0.59
        ws.page_margins.bottom = 0.59

        ws.header_footer.center_header = "Seite &P/&N"
        if header_right_text:
            ws.header_footer.right_header = header_right_text
    except Exception:
        pass

    ws.freeze_panes = f"A{header_row_idx + 1}"
    wb.save(path)
